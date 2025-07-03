from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def generate_pdf_report(resultados, proyecto_id):
    """Genera un reporte PDF con los resultados de los cálculos."""
    # Crear directorio reports si no existe
    if not os.path.exists('reports'):
        os.makedirs('reports')
    
    # Nombre del archivo
    filename = f'reports/proyecto_{proyecto_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    # Crear documento
    doc = SimpleDocTemplate(
        filename,
        pagesize=letter,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Contenido
    elements = []
    
    # Título
    elements.append(Paragraph(f'Reporte de Cálculos Eléctricos', title_style))
    elements.append(Paragraph(f'Proyecto {proyecto_id}', subtitle_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}', normal_style))
    elements.append(Spacer(1, 24))
    
    # Tabla de resultados
    data = [
        ['ID', 'Descripción', 'Corriente (A)', 'Calibre', '%Caída', 'Canalización', 'Estado']
    ]
    
    for r in resultados:
        estado = 'OK' if r['Caida_Tension'] <= 3 else 'Revisar' if r['Caida_Tension'] <= 5 else 'Fuera de rango'
        data.append([
            r['ID_Equipo'],
            r['Descripcion'],
            f"{r['Corriente']:.2f}",
            r['Calibre'],
            f"{r['Caida_Tension']:.2f}%",
            r['Canalizacion'],
            estado
        ])
    
    # Estilo de tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Generar PDF
    doc.build(elements)
    
    return filename

def generate_excel_report(resultados, proyecto_id):
    """Genera un reporte Excel con los resultados de los cálculos."""
    # Crear directorio reports si no existe
    if not os.path.exists('reports'):
        os.makedirs('reports')
    
    # Nombre del archivo
    filename = f'reports/proyecto_{proyecto_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resultados'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Descripción', 'Corriente (A)', 'Calibre', '%Caída', 'Canalización', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row, r in enumerate(resultados, 2):
        estado = 'OK' if r['Caida_Tension'] <= 3 else 'Revisar' if r['Caida_Tension'] <= 5 else 'Fuera de rango'
        data = [
            r['ID_Equipo'],
            r['Descripcion'],
            round(r['Corriente'], 2),
            r['Calibre'],
            round(r['Caida_Tension'], 2),
            r['Canalizacion'],
            estado
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            # Colorear estado
            if col == 7:  # columna de estado
                if value == 'OK':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == 'Revisar':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar archivo
    wb.save(filename)
    
    return filename
